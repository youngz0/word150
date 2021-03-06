# coding=utf-8
import json,os
import numpy as np
import pandas as pd
from glob import glob
from PIL import Image
from pathlib import Path
import nudged
from tqdm import tqdm
import seaborn as sns
import cv2
from dim_utils_v3_1 import *
import scipy
from scipy.stats import *
import matplotlib as mpl
import matplotlib.pyplot as plt
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Color, Font, Alignment,PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
mpl.use('Agg')
zhfont = mpl.font_manager.FontProperties(fname = '/usr/share/fonts/truetype/arphic-gkai00mp/gkai00mp.ttf')
# # # # # =======================================================================================================
'''# 把标注的json格式转成打分时需要的格式'''
def cnvjsforscore(whichword):
    df_cfg=pd.read_excel('./prefilecfg/id_word_title.xlsx')
    idex = list(df_cfg[df_cfg['word'] == whichword].index.values)[0]
    whichwd = df_cfg.loc[idex,'pinyin']
    pth1 = os.path.join('./words_150/',whichwd,'ai_data/')
    jssavepth = os.path.join('./words_150/',whichwd+'_givescore','convstrucjson')
    if not os.path.exists(jssavepth):os.makedirs(jssavepth)
    jsfllst = glob(pth1 + '*' + '.json')
    if len(glob(jssavepth+'/'+'*'+'.json'))  != len(jsfllst):
        for i in jsfllst:
            te = {}
            te['images'] = '2/2'
            te['annotations'] = read_points(i)
            te['results'] = [1]
            with open(jssavepth+'/'+ os.path.split(i)[1] , 'w') as f:
                json.dump(te, f,indent=4)
# # # # # =======================================================================================================
def fs(frpt,secpt,thirdpara,fourthpara,fifthpara):
    # # 输入的是list  //// 'load_data('+','.join([frpt,secpt,thirdpara,fourthpara,fifthpara])+')'
    singleload_data = list(map(lambda f1,s1,t1,f2,f3:'load_data('+','.join([f1,s1,t1,f2,f3])+')', frpt,secpt,thirdpara,fourthpara,fifthpara))
    return ','.join(singleload_data)
# # # # # =======================================================================================================
'''# 处理xlsx表格中 根据first_point、second_point 不同的输入方式来写 如何load_data'''
def firstsec_point(funcnm,tmdf):
    # if funcnm in ['cross_p_H','cross_p_V']:
    if ';' in list(tmdf['first_point'])[0]:
        a1111 = []
        for kkkk in tmdf.index:
            llst = tmdf.loc[kkkk,'first_point'].split(';')[1]
            az = ','.join(list(map(lambda a2,b2,c2,d2,e2:"load_data(" + str(a2) + ',' + str(b2) + ','+c2+ ',' +d2+ ',' +e2+")",\
                tmdf.loc[kkkk,'first_point'].split(';')[0].split(','),tmdf.loc[kkkk,'second_point'].split(','),\
                    [tmdf.loc[kkkk,'thirdpara']]*len(tmdf.loc[kkkk,'first_point'].split(',')),\
                    [tmdf.loc[kkkk,'fourthpara']]*len(tmdf.loc[kkkk,'first_point'].split(',')),\
                    [tmdf.loc[kkkk,'fifthpara']]*len(tmdf.loc[kkkk,'first_point'].split(',')))))
            b1111 = "['"+tmdf.loc[kkkk,'dimensionname']+"',"+az+",'"+tmdf.loc[kkkk,'meaning']+"']"
            b1111 = "['"+tmdf.loc[kkkk,'dimensionname']+"',"+az+",["+llst+"],'"+tmdf.loc[kkkk,'meaning']+"']"
            a1111.append(b1111)
        # funcnmtransf = 'evaluate_'+funcnm
        outres = '['+','.join(a1111)+']'
    # elif funcnm in ['double_aware','double_sleep','nline_parallel']:
    elif ',' in list(tmdf['first_point'])[0] and ';' not in list(tmdf['first_point'])[0]:
        a1111 = []
        for kkkk in tmdf.index:
            # 暂时不想起名字
            az = ','.join(list(map(lambda a2,b2,c2,d2,e2:"load_data(" + str(a2) + ',' + str(b2) + ','+c2+ ',' +d2+ ',' +e2+")",\
                tmdf.loc[kkkk,'first_point'].split(','),tmdf.loc[kkkk,'second_point'].split(','),\
                    [tmdf.loc[kkkk,'thirdpara']]*len(tmdf.loc[kkkk,'first_point'].split(',')),\
                    [tmdf.loc[kkkk,'fourthpara']]*len(tmdf.loc[kkkk,'first_point'].split(',')),\
                    [tmdf.loc[kkkk,'fifthpara']]*len(tmdf.loc[kkkk,'first_point'].split(',')))))
            b1111 = "['"+tmdf.loc[kkkk,'dimensionname']+"',"+az+",'"+tmdf.loc[kkkk,'meaning']+"']"
            a1111.append(b1111)
        outres = '['+','.join(a1111)+']'
    elif list(tmdf['first_point'])[0] == 'AA':
        a1111 = []
        for kkkk in tmdf.index:
            az = '['+tmdf.loc[kkkk,'second_point']+']'
            b1111 = "'"+tmdf.loc[kkkk,'dimensionname']+"',"+az+",'"+tmdf.loc[kkkk,'meaning']+"'"
            a1111.append(b1111)            
        outres = '['+','.join(a1111)+']'
    else:
        outres = '[' + ','.join(list(map(lambda a1,b1,c1,d1,e1,f1,g1:"'"+a1+"',load_data(" + str(int(float(b1))) + ',' + str(int(float(c1))) + ','+d1+ ',' +e1+ ',' +f1+"),'"+g1+"'"  ,\
            list(tmdf['dimensionname']),list(tmdf['first_point']),list(tmdf['second_point']),list(tmdf['thirdpara']),\
            list(tmdf['fourthpara']),list(tmdf['fifthpara']),list(tmdf['meaning']) ))) + ']' 
    return outres
# # # # # =======================================================================================================
'''# 读取json文件 input json文件名；   返回 json内的数据'''
def rdjson(flnm):
    with open(flnm) as js:
        data = json.load(js)
    return data
# # # # # =======================================================================================================
'''# 给范字图片加上红点 input：范字文件夹路径'''
def step2_show_point_singlwd(whichword):
    save_path = whichword
    jsons = glob(whichword +'/' +'*'+whichword.split('/')[-2]+'*'+'.json')
    redpointpic = glob(whichword +'/' +'*'+'label'+'*'+whichword.split('/')[-2]+'*')
    if redpointpic == []:
        for json_path in tqdm(jsons):
            img_path=json_path[:-5]+'.png'
            img=Image.open(img_path)
            plt.imshow(img)
            jsonData=json.load(open(json_path,'r'))
            points_ = jsonData['shapes']
            p_l = []
            for p in points_:
                p_l.append([p['points'][0][0], p['points'][0][1]])
            coord = p_l
            for idx,c in enumerate(coord):
                plt.plot(c[0],c[1],'r.')
                plt.text(c[0],c[1],str(idx+1),fontsize= 8 ,weight = "light",color='black')
            plt.savefig(os.path.join(save_path,'label_' + img_path.split('/')[-1].split('.')[0] + '.png'))
            plt.clf()
# # # # # =======================================================================================================
'''# 从200179读取相应json生成csv表格,默认删掉b开头的;  input 汉字； 保存文件在本地pre_csv；  返回生成的csv路径'''
# def getjs2csvpre(whichword,*arg):
def step0_getjs2csvpre(whichword):
    folderpath = './prefilecfg/200179/'
    templatecsvnm = './prefilecfg/template.xlsx'
    # df_cfg=pd.read_csv('./prefilecfg/id_word_title.csv')
    df_cfg=pd.read_excel('./prefilecfg/id_word_title.xlsx')
    idex = list(df_cfg[df_cfg['word'] == whichword].index.values)[0]
    '''df_cfg.loc[idex,'idlenfour']    df_cfg.loc[idex,'title']    df_cfg.loc[idex,'word']    df_cfg.loc[idex,'pinyin']'''
    flnm = '_'.join( [str(df_cfg.loc[idex,'idlenfour']) , df_cfg.loc[idex,'title']] )
    jsfllist = glob(folderpath+'*'+flnm+'*'+'.json') 
    if jsfllist != []:
    # try:
        te = rdjson(jsfllist[0])
        tes = []
        data = te['data']
        title = data['word_name']
        temp = data['review']
        lstkey_pre = list(temp.keys())
        lstkey=list(set(lstkey_pre).difference(set(['b1zfgy', 'b2bfgy', 'b2zfgy', 'b3bhxfbf','tsjzf'])))
        lstkey.sort()
        for i in lstkey:
            klst = list(set(temp[i].keys())-set(['name', 'serial_num']))
            klst.sort()
            for j in klst:
                keyl = list(set(temp[i][j].keys())-set(['gif_file', 'gif_name', 'name', 'serial_num']))
                keyl.sort()
                for k in keyl:
                    tes.append([i+'__'+j+'__'+k,temp[i][j][k]['name'],temp[i][j]['name']])
        df = pd.read_excel(templatecsvnm,sep=',',header=None,index_col=None)    
        colnm = pd.read_excel(templatecsvnm,sep=',',skiprows=24,nrows=0,header=0)
        wdpinyin = df_cfg.loc[idex,'pinyin']
        # 'smbpath	/run/user/1000/gvfs/smb-share:server=192.168.60.181,share=share/sfq/word150/'
        df.loc[1,1] = whichword;  df.loc[2,1] = wdpinyin
        df.loc[4,1] = df.loc[4,1] + str(df_cfg.loc[idex,'idlenfour'])+'“'+df_cfg.loc[idex,'word']+'”字'#+'/ai_data'
        '''保存文件命名'''
        savepath = './pre_csv/' + wdpinyin + '-.xlsx'
        # df.to_csv(savepath,mode='w',index=False,header=False) #csv分两次写入，已改为直接存xlsx
        res = pd.DataFrame(tes,columns=['dimensionname','meaning','stroke'])
        res['thirdpara'] = 'p_Transed_l'; res['fourthpara'] = 'eg_p_l'; res['fifthpara'] = 'p_l'; res['delete_or_not'] = 0
        df2 = pd.merge(colnm,res,how='outer')
        df3 = df2[colnm.columns]
        '''只是初步根据条件填一部分，可能会有错'''
        pras1 = '夸'+whichword+'字较好'; pras2 = '鼓励'+whichword+'字写好'
        df3.loc[:,'SORTord'] = df3.index
        dntkindex = np.where(np.array(df3['meaning']) == '1.课节不对应提交')[0][0]
        df3.loc[dntkindex:,'delete_or_not'] = [9]*len(df3.loc[dntkindex:,'delete_or_not'])
        for idx in df3.index:
            # ----------------------------------------------------------------------------------------------------------
            '''## 先这么写有时间 改为 用字典的  // 可能规律找的不准，最后生成的表格不一定完全准确'''
            if df3.loc[idx,'dimensionname'].startswith('a2'):
                df3.loc[idx,'COLS_NMB'] = 1
            if df3.loc[idx,'dimensionname'].startswith('a3'):
                df3.loc[idx,'COLS_NMB'] = 2
            if df3.loc[idx,'dimensionname'].startswith('a4'):
                df3.loc[idx,'COLS_NMB'] = 3
            if df3.loc[idx,'dimensionname'].startswith('a5'):
                df3.loc[idx,'COLS_NMB'] = 4
            # ----------------------------------------------------------------------------------------------------------
            if  '斜点太圆' in df3.loc[idx,'meaning'] or '太重' in df3.loc[idx,'meaning']\
                or '出锋' in df3.loc[idx,'meaning'] or '颤笔' in df3.loc[idx,'meaning']\
                or '收笔顿笔回勾' in df3.loc[idx,'meaning'] or '收笔回勾' in df3.loc[idx,'meaning']\
                or '收笔顿笔回钩' in df3.loc[idx,'meaning'] or '收笔写成回钩' in df3.loc[idx,'meaning']\
                or '起笔位置不准' in df3.loc[idx,'meaning'] or '太方' in df3.loc[idx,'meaning']\
                or '行笔不顺直' in df3.loc[idx,'meaning'] or '垂点太圆' in df3.loc[idx,'meaning']\
                or '衔接不自然' in df3.loc[idx,'meaning']:
                '''*******************'''
                df3.loc[idx,'delete_or_not'] = 2
            # ----------------------------------------------------------------------------------------------------------
            if df3.loc[idx,'stroke'] + '较好' in  df3.loc[idx,'meaning'] or df3.loc[idx,'stroke'] + '行笔较好' in  df3.loc[idx,'meaning']:
                df3.loc[idx,'delete_or_not'] = 2
            # ----------------------------------------------------------------------------------------------------------
            if  pras1 in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'a1k1'; df3.loc[idx,'whichrowpnb'] = 'a1k1'
                df3.loc[idx,'first_point'] = 0; df3.loc[idx,'second_point'] = 0; df3.loc[idx,'delete_or_not'] = 4
            # ----------------------------------------------------------------------------------------------------------
            if  pras2 in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'a1k2'; df3.loc[idx,'whichrowpnb'] = 'a1k2'
                df3.loc[idx,'first_point'] = 0; df3.loc[idx,'second_point'] = 0;df3.loc[idx,'delete_or_not'] = 4
            # ----------------------------------------------------------------------------------------------------------
            if '有起收意识' in df3.loc[idx,'meaning'] or '有起收笔意识' in df3.loc[idx,'meaning']:
                df3.loc[idx,'CONS_ROWNMB'] = 1
                df3.loc[idx,'funcname'] = 'double_aware'
                df3.loc[idx,'whichrowpnb'] = 'positive'
            # ----------------------------------------------------------------------------------------------------------
            if '起收无顿笔' in df3.loc[idx,'meaning'] or '起收没有顿笔' in df3.loc[idx,'meaning'] or '起收笔没有' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'double_sleep'
                df3.loc[idx,'whichrowpnb'] = 'negative'
            # ----------------------------------------------------------------------------------------------------------
            if '起笔无顿笔' in df3.loc[idx,'meaning'] or '收笔无顿笔' in df3.loc[idx,'meaning']\
                or '起笔没有顿笔' in df3.loc[idx,'meaning'] or '收笔没有顿笔' in df3.loc[idx,'meaning']\
                or '转折无顿笔' in df3.loc[idx,'meaning'] or '转折处无顿笔' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'len_single'
                df3.loc[idx,'whichrowpnb'] = 'negative'
            # ----------------------------------------------------------------------------------------------------------
            if '布白较好' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'word_pos'
                df3.loc[idx,'whichrowpnb'] = 'negative'
                df3.loc[idx,'thirdpara'] = 'p_l'
            # ----------------------------------------------------------------------------------------------------------
            if '位置不居中' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'word_pos'
                df3.loc[idx,'whichrowpnb'] = 'positive'
                df3.loc[idx,'thirdpara'] = 'p_l'
            # ----------------------------------------------------------------------------------------------------------
            if '字形太大' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'len'
                df3.loc[idx,'whichrowpnb'] = 'positive'
                df3.loc[idx,'thirdpara'] = 'p_l'
            # ----------------------------------------------------------------------------------------------------------
            if '字形太小' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'len'
                df3.loc[idx,'whichrowpnb'] = 'negative'
                df3.loc[idx,'thirdpara'] = 'p_l'
            # ----------------------------------------------------------------------------------------------------------
            if '太长' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'len'
                df3.loc[idx,'whichrowpnb'] = 'positive'
            # ----------------------------------------------------------------------------------------------------------
            if '太短' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'len'
                df3.loc[idx,'whichrowpnb'] = 'negative'
            # ----------------------------------------------------------------------------------------------------------
            # if ('长短' in df3.loc[idx,'meaning'] or '长度' in df3.loc[idx,'meaning']) and '好' in df3.loc[idx,'meaning']:
            if '长' in df3.loc[idx,'meaning']  and '好' in df3.loc[idx,'meaning']:
                df3.loc[idx,'CONS_ROWNMB'] = 1                
                df3.loc[idx,'funcname'] = 'len'
                df3.loc[idx,'whichrowpnb'] = 'badlabs'
            # ----------------------------------------------------------------------------------------------------------
            if '长短' in df3.loc[idx,'meaning'] and '不准' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'len'
                df3.loc[idx,'whichrowpnb'] = 'ele'                            
            # ----------------------------------------------------------------------------------------------------------
            if '偏平' in df3.loc[idx,'meaning'] or '行笔方向太平' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'slope_pp'
                df3.loc[idx,'whichrowpnb'] = 'negative'
            # ----------------------------------------------------------------------------------------------------------
            if '偏竖直' in df3.loc[idx,'meaning'] or '行笔方向太竖直' in df3.loc[idx,'meaning'] or '行笔方向太垂直' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'slope_cz'
                df3.loc[idx,'whichrowpnb'] = 'badlabs'
            # ----------------------------------------------------------------------------------------------------------
            if '倾斜角度不准' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'slope_tx'
                df3.loc[idx,'whichrowpnb'] = 'ele'
            # ----------------------------------------------------------------------------------------------------------
            if '向右倾斜' in df3.loc[idx,'meaning'] or '行笔方向偏右' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'slope_tx'
                # df3.loc[idx,'whichrowpnb'] = 'positive'
            # ----------------------------------------------------------------------------------------------------------
            if '向左倾斜' in df3.loc[idx,'meaning'] or '行笔方向偏左' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'slope_tx'
                # df3.loc[idx,'whichrowpnb'] = 'negative'
            # ----------------------------------------------------------------------------------------------------------
            if '斜度过大' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'slope_tx'
            # ----------------------------------------------------------------------------------------------------------
            if '起笔方向不准' in df3.loc[idx,'meaning'] or '收笔方向不准' in df3.loc[idx,'meaning'] \
                or '顿笔方向不准' in df3.loc[idx,'meaning'] or '出钩方向不准' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'angle_3points'
                df3.loc[idx,'whichrowpnb'] = 'ele'
            # ----------------------------------------------------------------------------------------------------------
            if '顿笔方向好' in df3.loc[idx,'meaning']:
                df3.loc[idx,'CONS_ROWNMB'] = 1
                df3.loc[idx,'funcname'] = 'angle_3points'
                df3.loc[idx,'whichrowpnb'] = 'badlabs'
            # ----------------------------------------------------------------------------------------------------------
            if '行笔太弯' in df3.loc[idx,'meaning']  and ('横' in df3.loc[idx,'stroke'] or '竖' in df3.loc[idx,'stroke']):
                df3.loc[idx,'funcname'] = 'curve'
                df3.loc[idx,'whichrowpnb'] = 'positive'
            # ----------------------------------------------------------------------------------------------------------
            if '行笔太弯' in df3.loc[idx,'meaning']  and ('撇' in df3.loc[idx,'stroke'] or '捺' in df3.loc[idx,'stroke'])  and '短撇'  not in df3.loc[idx,'stroke']:
                df3.loc[idx,'funcname'] = 'radian'
                # df3.loc[idx,'whichrowpnb'] = 'positive'
            # ----------------------------------------------------------------------------------------------------------
            if '左高右低（向下倾斜）' in df3.loc[idx,'meaning']  and '横' in df3.loc[idx,'stroke']:
                df3.loc[idx,'funcname'] = 'slope_tx'
                df3.loc[idx,'whichrowpnb'] = 'negative'
            # ----------------------------------------------------------------------------------------------------------
            if '夹角太大' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'angle_2lines'
                df3.loc[idx,'whichrowpnb'] = 'positive'
            # ----------------------------------------------------------------------------------------------------------
            if '夹角太小' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'angle_2lines'
                df3.loc[idx,'whichrowpnb'] = 'negative'
            # ----------------------------------------------------------------------------------------------------------
            if '太长' in df3.loc[idx,'meaning'] and '太短' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = '2parts_size_ratio'
            # ----------------------------------------------------------------------------------------------------------
            if '横长撇短不准' in df3.loc[idx,'meaning'] or '横短撇长不准' in df3.loc[idx,'meaning'] or '竖长横短' in df3.loc[idx,'meaning']\
                or '竖短横长' in df3.loc[idx,'meaning'] or '横长竖短' in df3.loc[idx,'meaning'] or '横短竖长' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = '2parts_size_ratio'
            # ----------------------------------------------------------------------------------------------------------
            if '行笔方向不准' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'slope_tx'
                df3.loc[idx,'whichrowpnb'] = 'ele'
            # ----------------------------------------------------------------------------------------------------------
            if '行笔方向好' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'slope_tx'
                df3.loc[idx,'whichrowpnb'] = 'badlabs'
            # ----------------------------------------------------------------------------------------------------------
            if '出钩太长' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'len_single'
                df3.loc[idx,'whichrowpnb'] = 'positive'
            # ----------------------------------------------------------------------------------------------------------
            if '无出钩' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = 'len_single'
                df3.loc[idx,'whichrowpnb'] = 'negative'
            # ----------------------------------------------------------------------------------------------------------
            if '左' in df3.loc[idx,'meaning'] and '高' in df3.loc[idx,'meaning'] and '右' in df3.loc[idx,'meaning']and '低' in df3.loc[idx,'meaning']:
                df3.loc[idx,'funcname'] = '2stroke_start_pos'
        
        # ---------------------------------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = wdpinyin
        fontstyl = Font(name='Arial', size=14)
        for row in dataframe_to_rows(df,index=False,header=False):
            ws.append(row)
        fstwsmax_column = ws.max_column
        
        for row2 in dataframe_to_rows(df3,index=False,header=False):
            ws.append(row2)
            if row2[10] > 2:
                cellrow = ws.iter_cols(min_col=0,max_col=fstwsmax_column+1,min_row=ws.max_row,max_row=ws.max_row)
                for cell in cellrow:
                    cell[0].fill = PatternFill("solid", fgColor="fbf9a2")

        ws.auto_filter.ref = 'A25:{}{}'.format(chr(97+ws.max_column-2),ws.max_row)
        for i in range(1,fstwsmax_column+1):
            for j in range(1,ws.max_row+1):
                cell = ws.cell(row=j,column=i)
                cell.font=Font(name='Arial', size=12)
        wb.save('./pre_csv/'+wdpinyin+'-.xlsx')
        wb.close()
        # ---------------------------------------------------        
        print('汉字：%s  拼音: %s   生成路径: %s   读取的json: %s '%(whichword,wdpinyin,savepath,jsfllist[0]))
    # except IndexError:
    #     print('cant find relevant json')
    return savepath
# # # # # =======================================================================================================
'''# 从远程smb  60.181 /sfq/word150 文件夹内copy 相应字的标注图片json、包括范字的 到本地words_150内，以字 的拼音为文件夹名
    input 该字初步生成的csv，只读取前7行;  返回 本地 范字所属文件夹'''
def step1_findincompatiblejson(whichword):
    # whichword  为 初步生成的csv文件路径，人工编辑之前的。  './pre_csv/ya_all_dimension.csv' 本函数只需要csv前7行内容
    df_1 = pd.read_excel(whichword,index_col=0,nrows=6,usecols=[0,1])
    lcalpth1 = os.path.join(df_1.loc['folderpath','value'],df_1.loc['word','value'],df_1.loc['data_path','value'])
    lcalpth2 = os.path.join(df_1.loc['folderpath','value'],df_1.loc['word','value'],df_1.loc['eg_json_path','value'])
    remotepth1 = os.path.join(df_1.loc['smbpath','value'],df_1.loc['data_path','value'])
    remotepth2 = os.path.join(df_1.loc['smbpath','value'],df_1.loc['eg_json_path','value'])
    picjsfllen1 = glob(lcalpth1+'/'+'*'+'.'+'*'); remtpicjsfllen1 = glob(remotepth1+'/'+'*'+'.'+'*')
    picjsfllen2 = glob(lcalpth2+'/'+'*'+'.'+'*'); remtpicjsfllen2 = glob(remotepth2+'/'+'*'+'.'+'*')
    fld = os.path.join(df_1.loc['folderpath','value'],df_1.loc['word','value'])
    if not os.path.exists(fld):os.makedirs(fld)
    if len(picjsfllen1) != len(remtpicjsfllen1):
        rmcmd1 = 'rm -r %s '%(lcalpth1)
        copycmd1 = 'cp -r %s %s'%(remotepth1,lcalpth1)
        print('ai_data 文件夹内数量不等 try to %s ,then copy from smb/181/'%rmcmd1)
        os.system(rmcmd1)
        os.system(copycmd1)
    if len(picjsfllen2) != len(remtpicjsfllen2):
        rmcmd2 = 'rm -r %s '%(lcalpth2)
        copycmd2 = 'cp -r %s %s'%(remotepth2,lcalpth2)
        os.system(rmcmd2)
        os.system(copycmd2)

    jsfllist = glob(lcalpth1 + '/' + '*' + '.json')
    te = []
    for i in jsfllist:
        tmp = rdjson(i)
        res = len(tmp['shapes'])
        te.append([i,res])
    res = pd.DataFrame(te,columns=['jsflnm','len'])
    eg = rdjson(glob(lcalpth2 + '/' + '*' + '.json')[0])
    egnm = len(eg['shapes'])
    unequallendf = res[res['len'] != egnm]
    if unequallendf.empty:
        print('{} {} ai_data 内有 {} 张图片/json'.format('-'*50,df_1.loc['hanzi','value'],len(glob(lcalpth1 + '/' + '*' + '.json'))))
    else:
        print("\033[37;41m\twarning: {} {} ai_data 内有 {} 个json文件的坐标点数量与范字不一致,请确认问题\033[0m".format('-'*25,df_1.loc['hanzi','value'],len(glob(lcalpth1 + '/' + '*' + '.json'))))
        print(unequallendf.loc[list(unequallendf.index),'jsflnm'])
    return lcalpth2
# # # # # =======================================================================================================
'''# 读取 json内的点坐标数据  input json文件名 '''
def read_points(filenm):
    data = open(filenm); data = json.load(data)
    ####-------标注的json和自动生成的json  key 不一样  ----------
    if 'annotations' in data.keys():
        points_l = data['annotations']
    else:
        points_ = data['shapes']
        points_l = []
        for p in points_:points_l.append(p['points'][0])
    return points_l
# # # # # =======================================================================================================
'''# 读取 人工编辑好的csv文件  input 汉字 '''
def getcsvinfo(whichword,is_save):
    df_cfg=pd.read_excel('./prefilecfg/id_word_title.xlsx')
    df = df_cfg.set_index(['word'])
    if is_save == 1:
        filenm = './pre_csv/{}.csv'.format(df.loc[whichword,'pinyin'])
        df_1 = pd.read_csv(filenm,sep=',',index_col=0,nrows=13,usecols=[0,1])
        df_2 = pd.read_csv(filenm,sep=',',skiprows=24,header=0,usecols=list(range(14)),converters={'first_point':str,'second_point':str})
    else:
        filenm = './pre_csv/{}.xlsx'.format(df.loc[whichword,'pinyin'])
        df_1 = pd.read_excel(filenm,index_col=0,nrows=13,usecols=[0,1])
        df_2 = pd.read_excel(filenm,skiprows=24,header=0,usecols=list(range(14)),converters={'first_point':str,'second_point':str})
    dict_1 = df_1.to_dict()[df_1.columns[0]]
    df_2 = df_2[df_2['delete_or_not'] < 5]
    return dict_1,df_2
# # # # # =======================================================================================================
'''# 生成word-read.json 文件;     生成word.json 文件 在results文件夹内  
    input 汉字  保存pinyin-read.json 、 pinyin.json  输出pinyin-read.json 路径'''
def convcsv2jsrd(whichword,is_save):
    temp,df2 = getcsvinfo(whichword,is_save)
    res = {}
    res['word'] = temp['word']
    res['unvpath'] =  os.path.join(temp['folderpath'], temp['word'])
    res['datapath'] = os.path.join(temp['folderpath'], temp['word'], temp['data_path'])
    res['egpath'] =   os.path.join(temp['folderpath'], temp['word'], temp['eg_json_path'])
    res['smbpath'] = temp['smbpath']
    res['savepath'] = os.path.join(temp['folderpath'], temp['word'], temp['save_path'])
    res['results'] =  os.path.join(temp['folderpath'], temp['word'], temp['jsresults'])
    res['figpath'] =  os.path.join(temp['folderpath'], temp['word'], temp['figpath'])
    res['finalrespath'] =  os.path.join('./','A_results','words_150',temp['word'], temp['jsresults'])
    res['meanstdcsv'] = temp['meanstdcsv'] + '_' + temp['word']
    res['csv_namepath'] = temp['csv_namepath'] + '_' + temp['word']
    res['json_para'] = temp['json_para'].split(',')
    res['result_para'] = temp['result_para'].split(',')
    egjsflname = glob(res['egpath'] + '/' + '*' + temp['word'] + '*'+'.json')[0]
    res['egpl'] = read_points(egjsflname)
    grp_a1 = df2[df2['delete_or_not'] == 4]
    df2 = df2[df2['delete_or_not'] < 2]
    res['evaluate_dim_l']  = list(map(lambda x,y:x+'__'+y,list(df2['dimensionname']),list(df2['whichrowpnb'])))
    df2['dimensionname'] = res['evaluate_dim_l']
    t = []
    for i in list(set(df2['funcname'])):
        if 1 in list(df2[df2['funcname'] == i]['vis_or_not']):
            t.append('1')
        else:
            t.append('0')
    dctk = list(map(lambda x:'evaluate_' + x + '_func' ,list(set(df2['funcname']))))    
    res['vis_or_not'] = dict(zip(dctk, t))
    df2 = df2.drop(columns=['vis_or_not', 'delete_or_not', 'whichrowpnb', 'COLS_NMB', 'CONS_ROWNMB'])
    df2['first_point'] = df2['first_point'].apply(str)  ##
    df2['second_point'] = df2['second_point'].apply(str)
    te = {}
    for i in list(set(df2['funcname'])):
        funcnmtransf = 'evaluate_' + i + '_func'
        var = {}
        tmp = df2[df2['funcname'] == i]
        for j in tmp.index:
            strokeinfo = dict(tmp.loc[j,:])
            k = strokeinfo['dimensionname']
            var[k] = strokeinfo
        te[funcnmtransf] = var
    res['rule_para'] = te
    ####--------------------生成文件夹
    for j in ['unvpath', 'datapath', 'egpath', 'savepath', 'results', 'figpath','finalrespath']:
        if not os.path.exists(res[j]):os.makedirs(res[j])
    rd = os.path.join(res['unvpath'], res['word'] + '-read.json')
    with open(rd, 'w') as f:
        json.dump(res, f,indent=4)
    # # # # # =======================================================================================================
    '''生成word.json'''
    # # key: dim_l
    v_0 = ''.join(list(map(lambda x :"self." + x + " = []\n", res['evaluate_dim_l'])))
    # # key: eg_p_l
    v_1 = read_points(egjsflname)
    # # key: evaluate_dim_l     v_2 = res['evaluate_dim_l']
    v_2 = list(df2['dimensionname'])
    # # key: name_dict
    v_3 = {}
    for i in v_2:
        v_3[i] = i.split('__')[:3]    
    for i in range(0,grp_a1.shape[0],1):
        v_3['a1k{}'.format(i+1)] = list(grp_a1['dimensionname'].values)[i].split('__')
    v_4 = {}
    print('计算 {} 字所用func： {}'.format(whichword,set(df2['funcname'])))
    for k in list(set(df2['funcname'])):
        tempdf = df2[df2['funcname'] == k]
        fspnt = firstsec_point(k,tempdf)
        funcnmtransf = 'evaluate_'+k
        v_4[funcnmtransf] = fspnt
    
    preprocess_data_v = ''
    tes = dict(zip(res['json_para'],[v_0,v_1,v_2,v_3,v_4,preprocess_data_v]))
    with open(res['results'] + '/'+ res['word'] + '.json', 'w') as f:
        json.dump(tes, f,indent=4)
    with open(res['finalrespath'] + '/'+ res['word'] + '.json', 'w') as f:
        json.dump(tes, f,indent=4)
    return rd
# # # # # =======================================================================================================
''' 生成pinyin_rules.json 文件   以 丫、一 为例 生成 ya_rules.json   yi_rules.json
    key 目前固定  ["CSV_PATH","DROPS","COLS_1","COLS_2","COLS_3","COLS_4","COLS_5","WTS_1","WTS_2","WTS_3",
                "WTS_4","WTS_5","CONS_1","CONS_2","CONS_3","CONS_4","PROS_1","CURVES","DETAILS","GLOBALPOSITIVE","GLOBALNEGATIVE","NAMES_DICT"]
    主要是给每一项key 赋值，决定value  .   可能需要修改    '''
def genjsrules(whichword,is_save):
    res,df2 = getcsvinfo(whichword,is_save)
    res['json_para'] = res['json_para'].split(',')
    res['result_para'] = res['result_para'].split(',')
    res['finalrespath'] =  os.path.join('./','A_results','words_150',res['word'], res['jsresults'])
    grp_a1 = df2[df2['delete_or_not'] == 4]
    df2 = df2[df2['delete_or_not'] < 2]
    res['evaluate_dim_l']  = list(map(lambda x,y:x+'__'+y,list(df2['dimensionname']),list(df2['whichrowpnb'])))
    df2['dimensionname'] = res['evaluate_dim_l']
    df3 = df2.copy()
    csvpath_v_lcal = res['folderpath'] + '/' +res['word'] + '/' + res['jsresults'] + '/' + 'dim_mean_std_' + res['word'] + '.csv'
    csvpath_v = '/app/ai_package/rules'+'/words_150' + '/' +res['word'] + '/' + res['jsresults'] + '/' + 'dim_mean_std_' + res['word'] + '.csv'
    drops_v = list(df3[df3['delete_or_not'] == 1]['dimensionname'])
    df3 = df3[df3['delete_or_not'] == 0]
    """ # 开始给COLS_1、COLS_2、COLS_3、COLS_4、COLS_5 , WTS_1、WTS_2、WTS_3、WTS_4、WTS_5 CONS_1~CONS_4等赋值
        # 目前是需要先在csv的COLS_NMB 定义好 分哪几类，用1、2、3、4、5等区分
        # 如果少于5，缺的为[],  可能要改    """
    for i in range(1,6,1):
        exec('COLS_{} = {}'.format(i, list(df3[df3['COLS_NMB'] == i]['dimensionname'].values)))
        exec('WTS_{} = {}'.format(i, [1]*len(locals()['COLS_{}'.format(i)]) ))
    # # # # # =======================================================================================================
    for i in range(1,5,1):
        a = df3[df3['COLS_NMB'] == i]
        exec('CONS_{} = {}'.format(i, list(a[a['CONS_ROWNMB'] != 1]['dimensionname'].values)))
    for i in range(1,5,1):
        a = df3[df3['COLS_NMB'] == i ]
        exec('PROS_{} = {}'.format(i, list(a[a['CONS_ROWNMB'] == 1]['dimensionname'].values)))
    DETAILS_v = []; CURVES_v = []
    for i in list(set(df3['funcname'])):
        if 'curve' in i  or 'slope' in i:
            CURVES_v = CURVES_v + list(df3[df3['funcname'] == i]['dimensionname'])
        if 'len' in i and 'single' not in i:
            DETAILS_v = DETAILS_v + list(df3[df3['funcname'] == i]['dimensionname'])
    NAMES_DICT_v = {}; GLOBALPOSITIVE_v = []; GLOBALNEGATIVE_v = []
    for i in list(df3['dimensionname']):
        NAMES_DICT_v[i] = i.split('__')[:3]
        if 'badlabs' not in i:
            GLOBALPOSITIVE_v.append(i)
    for i in range(0,grp_a1.shape[0],1):
        NAMES_DICT_v['a1k{}'.format(i+1)] = list(grp_a1['dimensionname'].values)[i].split('__')
    jsres = dict(zip(res['result_para'],\
            [csvpath_v,drops_v,\
            locals()[res['result_para'][2]],locals()[res['result_para'][3]],locals()[res['result_para'][4]],\
            locals()[res['result_para'][5]],locals()[res['result_para'][6]],\
            locals()[res['result_para'][7]],locals()[res['result_para'][8]],locals()[res['result_para'][9]],\
            locals()[res['result_para'][10]],locals()[res['result_para'][11]],\
            locals()[res['result_para'][12]],locals()[res['result_para'][13]],locals()[res['result_para'][14]],\
            locals()[res['result_para'][15]],\
            locals()[res['result_para'][16]],locals()[res['result_para'][17]],locals()[res['result_para'][18]],\
            locals()[res['result_para'][19]],\
            CURVES_v,DETAILS_v,GLOBALPOSITIVE_v,GLOBALNEGATIVE_v,NAMES_DICT_v]))
    jsres_lcal = jsres.copy() 
    jsres_lcal['CSV_PATH'] = csvpath_v_lcal
    outp = os.path.join(res['folderpath'],res['word'],res['jsresults'],res['word'] + '_rules.json')
    print('word_rules.json save path : {}'.format(outp))
    '''# 生成两份结果一份存在 A_results里面只包含要提交的文件，且pinyin_rules.json中CSV_PATH路径是提交后会用到的，另一份CSV_PATH路径是Step4需要用到'''
    with open(outp, 'w') as f:
        json.dump(jsres_lcal, f,indent=4)
    with open(res['finalrespath'] + '/'+ res['word'] + '_rules.json', 'w') as f:
        json.dump(jsres, f,indent=4)
# # # # # =======================================================================================================
def fnd_maxmin(p_l_input):
    x1 = np.min([i[0] for i in p_l_input]); x2 = np.max([i[0] for i in p_l_input])
    y1 = np.min([i[1] for i in p_l_input]); y2 = np.max([i[1] for i in p_l_input])
    return x1,x2,y1,y2
# # # # # =======================================================================================================
def final_run_ver(whichworld,is_save):
    con = convcsv2jsrd(whichworld,is_save)
    genjsrules(whichworld,is_save)
    word = review_word(con,is_save)
    word.cal_score()
# # # # # =======================================================================================================
def figrename8(whichword,nmb,savpath):
    fltoread = glob('./pre_csv/'+'*'+whichword+'.csv')
    if fltoread != []:
        df = pd.read_csv(fltoread[0],sep=',',skiprows=24,header=0)
        df['dimensionname'] =  list(map(lambda x,y:x+'__'+y,list(df['dimensionname']),list(df['whichrowpnb'])))
        dm = list(df[df['delete_or_not'] != 4]['dimensionname'])[0]
        print('dim: %s   含义： %s   所用函数： %s'%(dm,list(df[df['dimensionname'] == dm]['meaning'])[0],list(df[df['dimensionname'] == dm]['funcname'])[0]))
        foldername = '_'.join([df.loc[0,'stroke'],list(df[df['dimensionname'] == dm]['stroke'])[0],list(df[df['dimensionname'] == dm]['meaning'])[0]])
    csv_path ='./words_150/'+whichword+'/'
    flt = glob(csv_path+'*'+ whichword+'.csv')
    if flt != []:
        df_2 = pd.read_csv(flt[0])
        df2 = df_2.sort_values(by=dm , ascending=False)
        df2 = df2.reset_index(drop=True)
    # srtest = 'cp -r  '+ csv_path+'vis_result  ' + csv_path+'vis_result_'+foldername
    srtest = 'cp -r  {}vis_result  {}'.format(csv_path,savpath+foldername)
    if not os.path.exists(csv_path+'vis_result_'+foldername):
        print(srtest)
        os.system(srtest)
    for i in list(df2[dm]):
        idxnm = list(df2[df2[dm] == i].index)[0]
        nam1 = os.path.join(savpath+foldername,list(df2[df2[dm] == i]['00_pic_name_l'])[0])
        nam2 = os.path.join(savpath+foldername,"%03d" % idxnm +'____'+str(i)+'____'+list(df2[df2[dm] == i]['00_pic_name_l'])[0])        
        try:
            os.rename(nam1,nam2)
        except:
            continue
    rmcmd2 = 'rm -r {}/*.*'.format(os.path.join('./words_150/',whichword,'vis_result'))
    os.system(rmcmd2)
# # # # # =======================================================================================================
# # # # 梳理需求：             1、计算，生成csv                                                           # # # #
# # # #                      2、根据规则生成json/yaml 文件                                               # # # #
# # # # # =======================================================================================================
class review_word():
    def __init__(self,jsonfl,is_save):
        js = open(jsonfl)
        data = json.load(js)                          # # 读取json文件， 以下7行分别是
        self.word = data['word']                      # # 某个字
        self.hanzi = data['smbpath'].split('“')[1].split('”')[0]
        self.unvpath = data['unvpath']+'/'
        self.data_path = data['datapath']+'/'             # # 该字对应的  存有标注图片和json文件 的 文件夹路径 目前是ai_data
        self.save_path = data['savepath']+'/'             # # 该字 可视化结果的保存路径
        self.eg_path = data['egpath']+'/'                 # # 范字的json文件路径
        self.csv_savepath = data['results']+'/'           # # 该字初次计算生成的csv文件保存路径
        self.figpath = data['figpath']+'/'
        self.csvname = data['csv_namepath'] #+ '_' + data['word']                             # # 该字初次计算生成的csv文件名
        self.meanstdcsvname = data['meanstdcsv']# + '_' + data['word']                        # # 该字初次计算生成的mean_std csv文件名
        self.result_para = data['result_para']                                               # # 生成json结果的 key
        self.json_par = data['json_para']                                                    # # 生成json_rule 结果的 key
        self.rule = data['rule_para']                                                        # # 该字对应的能通用的规则rule
        self.visornot = data['vis_or_not']
        self.pic_name_l = []
        self.score_dict = {}
        self.finalrespath = data['finalrespath']
        evaluate_dim_l = data['evaluate_dim_l']    # 需要评价的，准备和书法老师协调
        self.evaluate_dim_l = evaluate_dim_l
        self.eg_p_l = data['egpl']
        self.is_save = is_save
        for i in evaluate_dim_l:
            setattr(self, i,[])
        self.getfilelist()
    # # # # =======================================================================================================
    # # 获取文件列表    输出json文件list、对应的图片文件名列表
    def getfilelist(self):
        js_file_list = []; pic_file_list = []
        for name in os.listdir(self.data_path):
            if name[-4:] == 'json':
                name_ = Path(self.data_path+name[:-4]+'jpg')
                pic_name = name[:-4]+'jpg' if name_.is_file() else name[:-4]+'png'
                js_file_list.append(name); pic_file_list.append(pic_name)
        self.js_list = js_file_list
        self.pic_name_list = pic_file_list
    # # # # =======================================================================================================
    # # setattr getattr  & load
    def ldvalue(self,whichword,whichrule,ipt,vis_or_not):
        attr = '{}_index_eva_{}'.format(whichword,whichrule.__name__)
        setattr(self, attr, whichrule(ipt,self.p_Transed_l,self.eg_p_l,self.p_l,vis_or_not))
        resultuple = getattr(self,attr)
        self.load_score(resultuple[0],resultuple[1])
    # # # # =======================================================================================================
    def filtdim(self,funcname):
        tmplst1 = list(set(self.rule.keys()).intersection(set(map(lambda x :'evaluate_' + x + '_func',funcname))))
        aa = list(map(lambda x :list(self.rule[x].keys()),tmplst1))
        tmkl1 = []
        [tmkl1.extend(i) for i in aa]
        return tmkl1
    # # # # =======================================================================================================
    def choose_dimension(self):
        print('-----1-----')
        # print('-----没想好，之前是人工筛选，每个字都要筛选，工作量太繁杂。准备和书法老师协调-----')
    # # ---------------------，先放着--------------------------------------------------
    def load_score(self,dim_list,score_list):
        for i in range(len(dim_list)):
            eval('self.'+dim_list[i]).append(score_list[i])
    def calculate_dim_score(self):
        eg_p_l = self.eg_p_l
        for js_name,pic_name in tqdm(zip(self.js_list, self.pic_name_list)):
            p_l = read_points(self.data_path +'/'+ js_name)
            self.pic_name_l.append(pic_name)
            trans = nudged.estimate(self.eg_p_l,p_l)
            trans_location = trans.get_translation()
            trans_scale = trans.get_scale()
            p_Transed_l = [[(i[0]-trans_location[0])/trans_scale,(i[1]-trans_location[1])/trans_scale] for i in p_l]
            self.p_l = p_l
            self.p_Transed_l = p_Transed_l
            # # # # # =======================================================================================================
            ''' 主要分为三类。
            类一：夸奖。 调用  evaluate_word_pos_func ，输入为 四个最大最小值。
            比较固定，先单列     （已经修改了evaluate_word_pos 的输入方式，合并在类三里）'''
            # x_1,x_2,y_1,y_2 = fnd_maxmin(p_l)
            # stroke_word_pos = ['praise__nxgzzj',[[x_1,y_1],[x_2,y_2]],'夸奖_1.能写格子中间']            
            # self.ldvalue(self.word,  evaluate_word_pos_func,  stroke_word_pos,0)
            # # # # # =======================================================================================================
            ''' 类二： 需要先行计算，输入参数 不能直接调用load_data。 主要调用evaluate_pos_func 
            还没想好怎么解决'''
            # ##preprocess_data
            # v_12 = vec(p_Transed_l[0],p_Transed_l[1])
            # eg_v_12 = vec(self.eg_p_l[0],self.eg_p_l[1])
          
            # v_78 = vec(p_Transed_l[6],p_Transed_l[7])
            # eg_v_78 = vec(self.eg_p_l[6],self.eg_p_l[7])  

            # len_12 = vec_len(v_12)
            # len_78 = vec_len(v_78)
            # eg_len_12 = vec_len(eg_v_12)
            # eg_len_78 = vec_len(eg_v_78)

            # chyz_wqsb = np.max([len_12/eg_len_12, len_78/eg_len_78])
            # ##preprocess_data
            # # # 难点，需要手动计算
            # stroke_pos = ['chyz_wqsb',[[chyz_wqsb],[0],['']],'无起收笔']
            # self.ldvalue(self.word,evaluate_pos_func,stroke_pos)
            # # # # # =======================================================================================================
            ''' 类三：输入参数可以直接调用load_data。可以按照
            stroke_angle = ['chyz_dbfxbz_qb',load_data(0,3,p_l,eg_p_l,p_l),'顿笔方向不准',
                            'chyz_dbfxbz_sb',load_data(5,8,p_l,eg_p_l,p_l),'顿笔方向不准'
                            ]
            stroke_curve = ['chyz_hblx',load_data(1,7,p_Transed_l,eg_p_l,p_l),'横波浪线']
            的格式，填入参数。'''
            # # k: whichrule
            for k in self.rule.keys():
                vis_or_not = int(self.visornot[k])
                keylist = list(self.rule[k].keys())
                i = keylist[0]
                tmdict = self.rule[k]
                res_cal = []
                if ',' not in self.rule[k][i]['first_point'] and 'AA' not in self.rule[k][i]['first_point']:
                    for ik in keylist:
                        firpnt = tmdict[ik]['first_point'].split(',')
                        loadstr = fs(firpnt,tmdict[ik]['second_point'].split(','),\
                            [tmdict[ik]['thirdpara']]*len(firpnt),[tmdict[ik]['fourthpara']]*len(firpnt),[tmdict[ik]['fifthpara']]*len(firpnt))
                        res_cal.extend([ik,eval(loadstr),tmdict[ik]['meaning']])
                elif ',' in self.rule[k][i]['first_point'] and ';' not in self.rule[k][i]['first_point']:
                    for ik in keylist:
                        firpnt = tmdict[ik]['first_point'].split(',')
                        loadstr = fs(firpnt,tmdict[ik]['second_point'].split(','),\
                            [tmdict[ik]['thirdpara']]*len(firpnt),[tmdict[ik]['fourthpara']]*len(firpnt),[tmdict[ik]['fifthpara']]*len(firpnt))
                        res_cal.append([ik]+list(eval(loadstr))+[tmdict[ik]['meaning']])
                elif ';' in self.rule[k][i]['first_point']:
                    for ik in keylist:
                        # firpnt = tmdict[ik]['first_point'].split(',')
                        firpnt = tmdict[ik]['first_point'].split(';')[0].split(',')
                        unoname = list(eval(tmdict[ik]['first_point'].split(';')[1]))
                        loadstr = fs(firpnt,tmdict[ik]['second_point'].split(','),\
                            [tmdict[ik]['thirdpara']]*len(firpnt),[tmdict[ik]['fourthpara']]*len(firpnt),[tmdict[ik]['fifthpara']]*len(firpnt))
                        res_cal.append([ik]+list(eval(loadstr))+[unoname]+[tmdict[ik]['meaning']])
                elif self.rule[k][i]['first_point'] == 'AA':
                    for ik in keylist:
                        loadstr = '[' + tmdict[ik]['second_point'] + ']'
                        teemp = [ik]
                        teemp.append(eval(loadstr))
                        teemp.extend([tmdict[ik]['meaning']])
                        res_cal.extend(teemp)
                else:
                    print('new type !')

                self.ldvalue(self.word,eval(k),res_cal,vis_or_not)
            # print('learning---------pls wait')
            if self.is_save == True:
                display(self.data_path,pic_name,self.save_path,self.is_save,p_l,eg_p_l)
                clearHis()
    
    def cal_score(self):
        def sort_dict_key(dict_):
            # print('delete')
            sorted_d = {}
            for i in sorted (dict_) :
                sorted_d[i] = dict_[i]
            return sorted_d

        self.calculate_dim_score()
        for i in self.evaluate_dim_l:
            self.score_dict[i] = eval('self.'+i)
        self.score_dict['00_pic_name_l'] = self.pic_name_l
        self.score_dict  = sort_dict_key(self.score_dict)
        dim_score_df = pd.DataFrame(self.score_dict)
        dim_score_df.to_csv(self.unvpath + self.csvname + '.csv')
        # # # # # =======================================================================================================
        df2 = dim_score_df.copy()
        df_chinese = pd.DataFrame(columns=dim_score_df.columns)
        for ki in self.rule.keys():
            keylist2 = list(self.rule[ki].keys());
            for t in keylist2:
                df_chinese.loc[0,t] = self.rule[ki][t]['stroke']+self.rule[ki][t]['meaning']
                # df2.loc[len(dim_score_df),t] = self.rule[ki][t]['stroke']+self.rule[ki][t]['meaning']
        df_chinese.to_csv(self.unvpath + self.csvname + '_chinese.csv')#,header = None)
        df2.to_csv(self.unvpath + self.csvname + '_chinese.csv',mode='a',header = None)
        # # # # # =======================================================================================================
        ## 本段是为了生成均值方差的csv文件  有时间再改
        df = dim_score_df.copy()
        df3 = pd.DataFrame(index = ['mean','std'])
        whichpnb_new = []
        for j in df.columns[1:]:
            data_col_whole = df[j]
            dealkl1 = self.filtdim(['double_aware','double_sleep'])
            dealkl2 = self.filtdim(['radian'])
            dealkl3 = self.filtdim(['len','len_single'])
            dealkl4 = self.filtdim(['angle_3points','2angle_3points'])
            # 以下几个排除的阈值 可以改
            if j in dealkl1:
                # data_columns = df[df[j] > 0.1][j]; data_columns_n = df[df[j] > 0.1][j].tolist()
                data_columns = df[(df[j] > 0.1) & (df[j] < 5)][j]; data_columns_n = df[(df[j] > 0.1) & (df[j] < 5)][j].tolist()
                print('del double_aware/double_sleep')
            elif j in dealkl2:
                data_columns = df[(df[j] > 0.05) & (df[j] < 5)][j]; data_columns_n = df[(df[j] > 0.05) & (df[j] < 5)][j].tolist()
                print('del radian')
            ## 计算长度，5倍以上的 极大  和  0.02倍的极小 都舍弃，这样最后的曲线 比较像
            elif j in dealkl3:
                data_columns = df[(df[j] > 0.02) & (df[j] < 5)][j]; data_columns_n = df[(df[j] > 0.02) & (df[j] < 5)][j].tolist()
                print('del len/len_single')
            elif j in dealkl4:
                data_columns = df[(df[j] > 0.1) & (df[j] < 5)][j]; data_columns_n = df[(df[j] > 0.1) & (df[j] < 5)][j].tolist()
                print('del angle_3points/2angle_3points')
            elif df[df[j] == 0].shape[0]/len(df) > 0.05:
                j_new = j.split('__')[-1] if 'zero' in j else j.split('__')[-1]+'_zero'
                print('{}:  {} 个计算结果值为0，总数 {}'.format(df_chinese.loc[0,j],df[df[j] == 0].shape[0],len(df)))
                data_columns = df[df[j] != 0][j]
                data_columns_n = df[df[j] != 0][j].tolist()
            else:
                j_new = j.split('__')[-1][:-5] if 'zero' in j else j.split('__')[-1]
                if 'zero' in j.split('__')[-1]:
                    print('0占比很小，但{}中有zero'.format(j))
                data_columns = df[j]
                data_columns_n = df[j].tolist()
            # whichpnb_new.append(j_new)
            if 'dddd123456dddwz' in j:
                a = data_columns_n - np.asarray([0]*len(data_columns_n))
                sd = np.sqrt(sum(a*a)/len(data_columns_n))
                data_zscore = (data_columns_n - np.asarray([0]*len(data_columns_n)))/sd
            else:
                data_zscore = (data_columns_n - np.asarray(data_columns_n).mean())/np.asarray(data_columns_n).std()
            # --------------
            dealklst = self.filtdim(['double_aware','double_sleep','len_single'])
            
            # if j in dealklst and 'wcg' in j:
            if '0123456789' in j:
                # df3.loc['mean',j] = np.asarray(data_columns_n).mean()
                # df3.loc['std',j] = np.asarray(data_columns_n).std()
                df3.loc['mean',j] = 1
                a = data_columns_n - np.asarray([1]*len(data_columns_n))
                sd = np.sqrt(sum(a*a)/len(data_columns_n))
                df3.loc['std',j] = sd
            else:            
                df3.loc['mean',j] = np.asarray(data_columns_n).mean()
                df3.loc['std',j] = np.asarray(data_columns_n).std()
            # 画图    save_flag > 1 可以改为别的判断
            if self.is_save > 1:
                if j == 'praise__nxgzzj':
                    figname=j
                else:
                    figname = self.hanzi + '__' + df_chinese.loc[0,j] + '__' +j.split('_')[-1]
                fig = plt.figure()
                ax1 = fig.add_subplot(1, 2, 1)
                print(figname)
                try:
                    sns.distplot(np.array(data_col_whole),bins=round(len(data_col_whole)/5),kde=True)
                    p_val = scipy.stats.normaltest(data_col_whole)[1]
                    if p_val > 0.05:
                        ax1.text(0,0.95,'p: {}'.format(p_val),fontsize=10, color = "g",transform = ax1.transAxes)
                    else:
                        ax1.text(0,0.95,'p: {}'.format(p_val),fontsize=10, color = "r",transform = ax1.transAxes)
                except RuntimeError:
                    continue
                sns.distplot(np.array(data_columns_n),bins=round(len(data_columns_n)/5),kde=True )
                ax1.set_title('原始值  {} 个 '.format(len(data_col_whole)),fontproperties=zhfont)
                # ------------------------------------------
                ax2 = fig.add_subplot(1, 2, 2)
                sns.distplot(data_zscore,bins=round(len(data_zscore)/5),kde=True,fit=norm)
                p_value = scipy.stats.normaltest(data_zscore)[1]
                if p_value > 0.05:
                    ax2.text(0,0.95,'p: {}'.format(p_value),fontsize=10, color = "g",transform = ax2.transAxes)
                else:
                    ax2.text(0,0.95,'p: {}'.format(p_value),fontsize=10, color = "r",transform = ax2.transAxes)
                ax2.set_title('-mean/std之后的值  {} 个'.format(len(data_zscore)),fontproperties=zhfont)
                plt.suptitle(figname,fontproperties=zhfont)
                try:
                    plt.savefig(self.figpath+figname, format='png', dpi=600, pad_inches = 0)
                except FileNotFoundError:
                    plt.savefig(self.figpath+figname.replace('/',''), format='png', dpi=600, pad_inches = 0)
                print('{}   {}'.format(figname,len(data_columns_n)))
                plt.cla()
                plt.close("all")
                print('normaltest data_columns_n :{}'.format(scipy.stats.normaltest(data_columns_n)))
                print('normaltest data_zscore    :{}'.format(scipy.stats.normaltest(data_zscore)))
                print('-------- testing {}  {}'.format('-'*50,figname))
            df3.to_csv(self.csv_savepath + self.meanstdcsvname + '.csv')
            df3.to_csv(self.finalrespath +'/'+ self.meanstdcsvname + '.csv')

''' 反馈计算不准时， 在pre_csv 文件夹内修改对应字的xlsx 文件，前23行不变（夸、鼓励）24行开始保留要检查的 笔画维度，本部分对24开始的每一行vis_or_not置1
    并按照计算结果对该维度 在 dim_score_pinyin.csv 文件中按照该列进行排序，对生成图片重命名  方便按计算结果顺序查看'''
'''会在pinyin.xlsx  中 vis_or_not 置1 的 每一行单独生成一个csv文件'''

def checkerror(hanzi):
    df_cfg=pd.read_excel('./prefilecfg/id_word_title.xlsx')
    df = df_cfg.set_index(['word'])
    pinyin = df.loc[hanzi,'pinyin']
    filenm = './pre_csv/'+pinyin+'.xlsx'
    df_2 = pd.read_excel(filenm,sep=',',skiprows=24,header=0)
    vis = df_2[df_2['vis_or_not'] == 1]
    for i in vis.index:
        df_1 = pd.read_excel(filenm,sep=',',nrows=25,header=None)
        df_1.to_csv('./pre_csv/' + pinyin + '.csv',mode='w',index=False,header=False)
        df3pre = df_2[df_2['delete_or_not'] == 4]
        df3pre.to_csv('./pre_csv/' + pinyin + '.csv',mode='a',index=False,header=False)
        df_3 = df_2.loc[[i],:]
        df_3.to_csv('./pre_csv/' + pinyin + '.csv',mode='a',index=False,header=False)
        final_run_ver(hanzi,1)
        figrename(pinyin,i)

# if __name__ == '__main__':
#     # lst = '岁 处 丫 芝 方 甘 内 氏 本 人'.split(' ')
#     # lst = '虫'.split(' ')
#     for par in lst:
#         checkerror(par)
flder = './instances_20210201/2021-01-28-19-18-43/'

whichword = '本'
df_cfg=pd.read_excel('./prefilecfg/id_word_title.xlsx')
df = df_cfg.set_index(['word'])
pinyin = df.loc[whichword,'pinyin']

revdf = pd.read_excel('./pre_xlsx/{}.xlsx'.format(whichword),index_col=0)

for i in revdf.index:
    picl = glob(flder + '*' + str(i) + '*')
    rmcmd1 = 'rm -r {}/*.*'.format(os.path.join('./words_150/',pinyin,'ai_data'))
    os.system(rmcmd1)
    # rmcmd2 = 'rm -r {}/*.*'.format(os.path.join('./words_150/',pinyin,'vis_result'))
    # os.system(rmcmd2)
    savpth = os.path.join('./words_150/',pinyin,str(i))
    for j in picl:
        pth = os.path.join('./words_150/',pinyin,str(i))
        if not os.path.exists(pth):os.makedirs(pth)
        cpcmd = 'cp {} {}'.format(j, os.path.join(pth,j.split('/')[-1]))
        os.system(cpcmd)

        cpcmd1 = 'cp {} {}'.format(j, os.path.join('./words_150/',pinyin,'ai_data'))
        os.system(cpcmd1)
        # print(1)
    
    filenm = './pre_csv/'+pinyin+'.xlsx'
    revcheckerrcsv = './pre_csv/'+pinyin+'.csv'
    df_2 = pd.read_excel(filenm,sep=',',skiprows=24,header=0)
    df_1 = pd.read_excel(filenm,sep=',',nrows=25,header=None)
    df_1.to_csv(revcheckerrcsv,mode='w',index=False,header=False)
    df3pre = df_2[df_2['delete_or_not'] == 4]
    df3pre.to_csv(revcheckerrcsv,mode='a',index=False,header=False)

    df2_ = df_2.set_index(['dimensionname'])
    df2_.loc[[revdf.loc[i,'dimensionname']],'vis_or_not'] = 1
    df_3 = df2_.loc[[revdf.loc[i,'dimensionname']]]
    df_3.to_csv(revcheckerrcsv,mode='a',index=True,header=False)

    final_run_ver(whichword,1)
    figrename8(pinyin,i,savpth)
    
    print(1)
